from openpyxl import load_workbook

from tkinter import *
from tkinter import filedialog


class Scroll_Frame(Frame):
    def __init__(self, root):

        Frame.__init__(self, root)
        self.canvas = Canvas(root, borderwidth=0)
        self.frame = Frame(self.canvas)
        self.vsb = Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.create_window((4,4), window=self.frame, anchor="nw",
                                  tags="self.frame")

        self.frame.bind("<Configure>", self.onFrameConfigure)

    def onFrameConfigure(self, event):
        '''Reset the scroll region to encompass the inner frame'''
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))


def browse_excel_file_call_back(import_text):
    filename = filedialog.askopenfilename(filetypes=(("Excel Files", "*.xlsx"),))
    import_text.delete(0, END)
    import_text.insert(0, filename)


def browse_text_file_call_back(data, headers, ordered_headers):
    filename = filedialog.askopenfilename(filetypes=(("Tab delimited", "*.txt"),))

    file = open(filename, mode='r')

    # Make list of headers needed
    out_headers = file.readline().split()
    input_headers_count = len(out_headers)

    for header in ordered_headers:
        if headers[header].get():
            out_headers.append(header)

    out_file = open('output.txt', 'w')

    out_file.write('\t'.join(out_headers) + '\n')

    # Throw away input headers
    out_headers = out_headers[input_headers_count:]

    all_lines = file.readlines()

    for line in all_lines:
        line_data = line.split()
        caws_id = int(line_data[0])

        out_file.write('\t'.join(line_data) + '\t')

        out_data = []
        if caws_id in data:
            row_data = data[caws_id]

            for header in out_headers:
                if header in row_data:
                    out_data.append(str(row_data[header]))
                else:
                    out_data.append('-')
        else:
            out_data = ['-'] * len(out_headers)

        out_file.write('\t'.join(out_data) + '\n')

    out_file.close()
    file.close()


def import_workbook(ret_value, cb_frame, filename):
    workbook = load_workbook(filename)
    ret_value['wb'] = workbook

    all_cols = list_all_cols(workbook.worksheets)
    all_cols_dict = get_all_headers_dict(all_cols)
    ret_value['ordered_headers'] = all_cols
    ret_value['cols_checked'] = all_cols_dict

    Label(cb_frame.frame, text="Which columns to include in the output").pack(side=TOP, anchor='w', padx=2, pady=2)

    for col in all_cols:
        all_cols_dict[col] = BooleanVar()
        Checkbutton(cb_frame.frame, text=col, variable=all_cols_dict[col]).pack(side=TOP, anchor='w', padx=2, pady=2)

    ret_value['data'] = extract_data(workbook)


def extract_data(workbook):
    result = {}

    for ws in workbook.worksheets:
        headers = get_headers_list(ws)

        if headers[0] != 'CAWS#':
            continue

        row = 3
        while True:
            if ws.cell(row=row, column=1).value is None or ws.cell(row=row, column=1).value == ' ':
                break

            caws_id = int(ws.cell(row=row, column=1).value)
            row_data = {}

            for i in range(len(headers)):
                row_data[headers[i]] = ws.cell(row=row, column=i+1).value

                if row_data[headers[i]] is None:
                    row_data[headers[i]] = '-'

            if caws_id in result:
                print("Duplicate caws id %d" % caws_id)
            else:
                result[caws_id] = row_data

            row += 1

    return result


def list_all_cols(w_sheets):
    cols = []

    for ws in w_sheets:
        col_headers = get_headers_list(ws)

        if col_headers[0] != "CAWS#":
            continue

        for col in col_headers:
            if col not in cols:
                cols.append(col)

    return cols


def get_headers_list(ws):
    result = []
    col_num = 1

    while True:
        if ws.cell(row=1, column=col_num).value is not None:
            col_name = ws.cell(row=1, column=col_num).value
        else:
            col_name = ws.cell(row=2, column=col_num).value

        if col_name is None or col_name == "":
            break

        result.append(col_name)
        col_num += 1

    return result


def get_all_headers_dict(headers_list):
    return {x: 0 for x in headers_list}


def debug(my_dict):
    for k, v in my_dict.items():
        print(k, v)


def main():
    root = Tk()
    root.title("CAWS Data Mapper")
    root.geometry("500x500")

    ret_value = {}
    # Button(root, text="Debug", command=lambda: debug(ret_value['data'])).pack(side=TOP)

    frame = Frame(root)
    frame.pack(side=TOP, fill=X)

    Label(frame, text="Import excel file").pack(side=LEFT, padx=2, pady=2)

    import_text = Entry(frame)
    import_text.pack(side=LEFT, padx=2, pady=2)

    Button(frame, text="Browse", command=lambda: browse_excel_file_call_back(import_text))\
        .pack(side=LEFT, padx=2, pady=2)

    Button(root, text="Import", command=lambda: import_workbook(ret_value, frame, import_text.get()))\
        .pack(side=TOP, anchor='w', padx=2, pady=2)

    frame = Frame(root)

    Label(frame, text="Input mapping file").pack(side=LEFT, padx=2, pady=2)

    Button(frame, text="Browse", command=lambda: browse_text_file_call_back(ret_value['data'],
                                                                            ret_value['cols_checked'],
                                                                            ret_value['ordered_headers']))\
        .pack(side=LEFT, padx=2, pady=2)

    frame.pack(side=BOTTOM, anchor='w')

    frame = Scroll_Frame(root)
    frame.pack(side=TOP, fill=X)

    root.mainloop()


main()
